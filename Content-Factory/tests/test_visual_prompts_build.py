from __future__ import annotations

import csv
from pathlib import Path

import pytest

from orchestrator.site_info_fields import (
    detect_gemini_refusal,
    extract_visual_prompt_full,
    parse_info_fields,
    validate_visual_prompt,
)
from orchestrator.site_visual_validate import (
    collect_story_visual_records,
    run_validate_site_info_visuals,
    write_visual_prompt_tables,
)
from orchestrator.visual_stage import TECHNICAL_PROMPT, VisualStory

LONG_VISUAL = (
    "A cinematic photorealistic portrait of a woman in a sunlit apartment, "
    "elegant wardrobe, soft natural light, shallow depth of field, detailed skin texture, "
    "35mm film grain, editorial fashion photography, moody atmosphere, rich colors, "
    "full body visible, tasteful non-explicit composition, high detail 8k."
)


def _workspace(tmp_path: Path, name: str, *, raw: str, info: str = "") -> Path:
    ws = tmp_path / "runs" / "stories" / name
    pipe = ws / "_pipeline"
    pipe.mkdir(parents=True)
    (pipe / "mapping.json").write_text(
        f'{{"story_id":"{name}","canonical_basename":"{name}"}}',
        encoding="utf-8",
    )
    (pipe / "site_info_raw.txt").write_text(raw, encoding="utf-8")
    out = tmp_path / "output" / "site" / name
    out.mkdir(parents=True)
    if info:
        (out / "info.txt").write_text(info, encoding="utf-8")
    return ws


def test_long_visual_prompt_not_truncated_in_extraction():
    info = f"Заголовок: Story One\nВизуал: {LONG_VISUAL}\nТип голоса: F\n"
    full = extract_visual_prompt_full(info)
    assert full == LONG_VISUAL
    assert not full.endswith("...")
    assert len(full) >= 150


def test_title_equals_canonical_is_invalid():
    ok, status, reason = validate_visual_prompt("Holiday Dream", canonical_basename="Holiday Dream")
    assert not ok
    assert status == "invalid"
    assert reason == "title_fallback_blocked"


def test_none_visual_is_invalid():
    ok, status, reason = validate_visual_prompt(None, canonical_basename="Story")
    assert not ok
    assert status == "missing"
    assert reason == "no_visual_prompt_found"


def test_ellipsis_suffix_is_invalid():
    truncated = "A" * 160 + "..."
    ok, status, reason = validate_visual_prompt(truncated, canonical_basename="Story")
    assert not ok
    assert status == "invalid"
    assert reason == "truncated_prompt_blocked"


def test_gemini_refusal_detected():
    assert detect_gemini_refusal("I'm sorry, I can't help with that request due to policy.")
    ok, status, reason = validate_visual_prompt(
        LONG_VISUAL,
        canonical_basename="Story",
        raw_text_for_refusal="I cannot help with that",
    )
    assert not ok
    assert reason == "gemini_refusal_or_policy_response"


def test_final_prompt_only_from_full_visual(tmp_path: Path):
    _workspace(
        tmp_path,
        "Good Story",
        raw=f"Заголовок: Good Story\nВизуал: {LONG_VISUAL}\nТип голоса: F\n",
    )
    records = collect_story_visual_records(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
    )
    assert len(records) == 1
    row = records[0]
    assert row.is_valid
    assert row.visual_prompt_full == LONG_VISUAL
    assert row.final_prompt == f"{TECHNICAL_PROMPT}; {LONG_VISUAL}"
    assert not row.final_prompt.endswith("...")


def test_missing_visual_goes_to_invalid_csv(tmp_path: Path):
    _workspace(
        tmp_path,
        "No Visual",
        raw="Заголовок: No Visual\nТип голоса: M\n",
    )
    export_dir = tmp_path / "visual"
    res = run_validate_site_info_visuals(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
        export_dir=export_dir,
    )
    assert res.report["no_visual_prompt_found"] >= 1
    assert res.report["valid_prompts"] == 0
    invalid_body = res.invalid_csv_path.read_text(encoding="utf-8-sig")
    assert "No Visual" in invalid_body
    valid_rows = list(csv.DictReader(res.valid_csv_path.open(encoding="utf-8-sig"), delimiter=";"))
    assert len(valid_rows) == 0


def test_site_info_raw_preferred_over_truncated_info(tmp_path: Path):
    _workspace(
        tmp_path,
        "324A",
        raw=f"Заголовок: 324A\nВизуал: {LONG_VISUAL}\n",
        info="Заголовок: 324A\nВизуал: short truncated piece...\n",
    )
    records = collect_story_visual_records(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
    )
    assert records[0].visual_prompt_full == LONG_VISUAL
    assert records[0].extraction_source == "site_info_raw.txt"


def test_missing_raw_invalid_even_if_info_has_visual(tmp_path: Path):
    out = tmp_path / "output" / "site" / "Only Info"
    out.mkdir(parents=True)
    (out / "info.txt").write_text(f"Визуал: {LONG_VISUAL}\n", encoding="utf-8")
    ws = tmp_path / "runs" / "stories" / "Only Info"
    ws.mkdir(parents=True)
    (ws / "_pipeline").mkdir()
    (ws / "_pipeline" / "mapping.json").write_text(
        '{"story_id":"Only Info","canonical_basename":"Only Info"}',
        encoding="utf-8",
    )
    records = collect_story_visual_records(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
    )
    assert records[0].failure_reason == "missing_raw"
    assert not records[0].is_valid


def test_valid_in_csv_invalid_excluded(tmp_path: Path):
    _workspace(tmp_path, "Valid One", raw=f"Визуал: {LONG_VISUAL}\n")
    _workspace(tmp_path, "Bad One", raw="Заголовок: Bad One\n")
    export_dir = tmp_path / "visual"
    res = run_validate_site_info_visuals(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
        export_dir=export_dir,
    )
    valid_names = {r["canonical_basename"] for r in csv.DictReader(res.valid_csv_path.open(encoding="utf-8-sig"), delimiter=";")}
    invalid_names = {r["canonical_basename"] for r in csv.DictReader(res.invalid_csv_path.open(encoding="utf-8-sig"), delimiter=";")}
    assert "Valid One" in valid_names
    assert "Bad One" in invalid_names
    assert "Bad One" not in valid_names
    assert res.xlsx_path is not None
    assert res.xlsx_path.is_file()


def test_parse_info_multiline_visual():
    info = "Заголовок: X\nВизуал: line one\ncontinuation line two\nЖанры: drama\n"
    fields = parse_info_fields(info)
    assert "line one\ncontinuation line two" in fields.get("visual", "")


def test_retryable_reasons_in_invalid_export(tmp_path: Path):
    _workspace(tmp_path, "Retry Me", raw="I can't help with that\n")
    export_dir = tmp_path / "visual"
    records = collect_story_visual_records(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
        export_dir=export_dir,
    )
    write_visual_prompt_tables(records, export_dir)
    inv = list(csv.DictReader((export_dir / "visual_prompts_invalid.csv").open(encoding="utf-8-sig"), delimiter=";"))
    assert inv[0]["failure_reason"] == "gemini_refusal_or_policy_response"


# ============================================================
# Новые тесты для "site_info валиден" контракта (см. site_info_fields).
# ============================================================


def test_is_site_info_workspace_valid_good(tmp_path: Path):
    from orchestrator.site_info_fields import is_site_info_workspace_valid

    raw = tmp_path / "site_info_raw.txt"
    raw.write_text(f"Заголовок: Good\nВизуал: {LONG_VISUAL}\n", encoding="utf-8")
    ok, reason, source = is_site_info_workspace_valid(
        canonical_basename="Good",
        raw_path=raw,
        json_path=None,
        info_path=None,
    )
    assert ok
    assert reason == ""
    assert source == "site_info_raw.txt"


def test_is_site_info_workspace_valid_refusal_blocks(tmp_path: Path):
    from orchestrator.site_info_fields import is_site_info_workspace_valid

    raw = tmp_path / "site_info_raw.txt"
    raw.write_text("I'm sorry, I can't help with that.\n", encoding="utf-8")
    ok, reason, _ = is_site_info_workspace_valid(
        canonical_basename="Bad",
        raw_path=raw,
        json_path=None,
        info_path=None,
    )
    assert not ok
    assert reason == "gemini_refusal_or_policy_response"


def test_is_site_info_workspace_valid_title_fallback_blocks(tmp_path: Path):
    from orchestrator.site_info_fields import is_site_info_workspace_valid

    raw = tmp_path / "site_info_raw.txt"
    raw.write_text("Заголовок: Holiday Dream\nВизуал: Holiday Dream\n", encoding="utf-8")
    ok, reason, _ = is_site_info_workspace_valid(
        canonical_basename="Holiday Dream",
        raw_path=raw,
        json_path=None,
        info_path=None,
    )
    assert not ok
    assert reason == "title_fallback_blocked"


def test_is_site_info_workspace_valid_missing_raw_blocks(tmp_path: Path):
    """info.txt с валидным `Визуал:` НЕ считается готовым без raw — у нас raw обязателен."""
    from orchestrator.site_info_fields import is_site_info_workspace_valid

    info = tmp_path / "info.txt"
    info.write_text(f"Визуал: {LONG_VISUAL}\n", encoding="utf-8")
    ok, reason, _ = is_site_info_workspace_valid(
        canonical_basename="X",
        raw_path=None,
        json_path=None,
        info_path=info,
    )
    assert not ok
    assert reason == "missing_raw"


def test_human_sync_copies_xlsx_and_csv(tmp_path: Path):
    """Эмулируем launch layout и проверяем, что validate копирует артефакты в human-папку."""
    # layout: tmp/Запуски/LAUNCH/10_Временные_файлы/legacy/runs/site/<id>-a
    launch = tmp_path / "Запуски" / "LAUNCH_TEST"
    runs_root = launch / "10_Временные_файлы" / "legacy" / "runs" / "site" / "RID-a"
    runs_root.mkdir(parents=True)

    stories_root = runs_root / "stories"
    stories_root.mkdir()
    output_site = launch / "10_Временные_файлы" / "legacy" / "output" / "site"
    output_site.mkdir(parents=True)

    ws = stories_root / "Good"
    pipe = ws / "_pipeline"
    pipe.mkdir(parents=True)
    (pipe / "mapping.json").write_text(
        '{"story_id":"Good","canonical_basename":"Good"}', encoding="utf-8"
    )
    (pipe / "site_info_raw.txt").write_text(
        f"Заголовок: Good\nВизуал: {LONG_VISUAL}\n", encoding="utf-8"
    )

    export_dir = runs_root / "visual"
    res = run_validate_site_info_visuals(
        runs_stories_dir=stories_root,
        output_site_dir=output_site,
        export_dir=export_dir,
        runs_root=runs_root,
    )
    assert res.ok
    expected_human = launch / "02_Сайт" / "03_Визуал_для_сайта"
    assert res.human_dir == expected_human
    assert res.human_xlsx_path is not None
    assert res.human_xlsx_path.is_file()
    assert (expected_human / "visual_prompts.csv").is_file()
    assert (expected_human / "visual_prompts_invalid.csv").is_file()
    assert (expected_human / "visual_prompts_build_report.json").is_file()


def test_human_sync_missing_path_with_require(tmp_path: Path):
    """Если launch не определяется и required=True, validate возвращает ok=False с понятной ошибкой."""
    stories_root = tmp_path / "stories"
    stories_root.mkdir()
    output_site = tmp_path / "output_site"
    output_site.mkdir()
    ws = stories_root / "Good"
    pipe = ws / "_pipeline"
    pipe.mkdir(parents=True)
    (pipe / "mapping.json").write_text(
        '{"story_id":"Good","canonical_basename":"Good"}', encoding="utf-8"
    )
    (pipe / "site_info_raw.txt").write_text(
        f"Заголовок: Good\nВизуал: {LONG_VISUAL}\n", encoding="utf-8"
    )
    export_dir = tmp_path / "visual"
    res = run_validate_site_info_visuals(
        runs_stories_dir=stories_root,
        output_site_dir=output_site,
        export_dir=export_dir,
        runs_root=tmp_path,
        require_human_dir=True,
    )
    assert not res.ok
    assert "human_visual_prompts_xlsx_path_not_found" in res.message


def test_summary_extra_in_xlsx(tmp_path: Path):
    """summary в xlsx должен содержать переданные extra-поля (retry метрики)."""
    from openpyxl import load_workbook

    _workspace(tmp_path, "Good", raw=f"Визуал: {LONG_VISUAL}\n")
    export_dir = tmp_path / "visual"
    records = collect_story_visual_records(
        runs_stories_dir=tmp_path / "runs" / "stories",
        output_site_dir=tmp_path / "output" / "site",
        export_dir=export_dir,
    )
    write_visual_prompt_tables(
        records,
        export_dir,
        summary_extra={
            "retry_status": "ok",
            "selected_gemini_profile": "user_data_1",
            "browser_launch_error": "",
        },
    )
    wb = load_workbook(export_dir / "visual_prompts.xlsx")
    sheet = wb["summary"]
    pairs = {row[0]: row[1] for row in sheet.iter_rows(values_only=True) if row and row[0]}
    assert pairs.get("retry_status") == "ok"
    assert pairs.get("selected_gemini_profile") == "user_data_1"
