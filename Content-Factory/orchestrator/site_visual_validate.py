from __future__ import annotations

import csv
import json
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from orchestrator.config import OrchestratorConfig
from orchestrator.site_info_fields import (
    RETRYABLE_FAILURE_REASONS,
    bump_report_counter,
    build_validation_report_counters,
    make_visual_prompt_preview,
    validate_story_visual_from_raw,
)
TECHNICAL_PROMPT = (
    "A photorealistic raw full-figure photograph. Natural, unposed lighting "
    "showing realistic skin texture. Shot on 35mm film, slight grain, high detail, 8k."
)

VALID_CSV_COLUMNS = [
    "canonical_basename",
    "story_id",
    "technical_prompt",
    "visual_prompt_full",
    "visual_prompt_preview",
    "final_prompt",
    "visual_prompt_status",
    "extraction_source",
    "story_workspace_path",
]

INVALID_CSV_COLUMNS = [
    "canonical_basename",
    "story_id",
    "visual_prompt_status",
    "failure_reason",
    "extraction_source",
    "raw_path",
    "story_workspace_path",
    "retry_count",
    "raw_excerpt_preview",
]

RETRY_STATE_FILE = "visual_prompts_retry_state.json"

# Human-launch layout: см. orchestrator/human_launch_layout.py.
HUMAN_LAUNCH_VISUAL_REL = ("02_Сайт", "03_Визуал_для_сайта")
HUMAN_LAUNCH_ROOT_MARKER = "10_Временные_файлы"

# Имя человекочитаемого Excel в launch (то же имя, чтобы пользователь привык).
HUMAN_VISUAL_XLSX_NAME = "visual_prompts.xlsx"
HUMAN_VISUAL_CSV_NAME = "visual_prompts.csv"
HUMAN_VISUAL_INVALID_CSV_NAME = "visual_prompts_invalid.csv"
HUMAN_VISUAL_REPORT_NAME = "visual_prompts_build_report.json"
HUMAN_VISUAL_RETRY_REPORT_NAME = "visual_prompts_retry_report.json"


@dataclass
class StoryVisualRecord:
    canonical_basename: str
    story_id: str
    story_workspace_path: Path
    output_story_dir: Path
    raw_path: Path | None
    raw_text: str
    site_info_json: dict[str, Any] | None
    info_text: str
    visual_prompt_full: str
    visual_prompt_preview: str
    visual_prompt_status: str
    failure_reason: str
    extraction_source: str
    final_prompt: str
    retry_count: int = 0
    raw_excerpt_preview: str = ""

    @property
    def is_valid(self) -> bool:
        return self.visual_prompt_status == "ok" and bool(self.visual_prompt_full.strip())


@dataclass
class SiteVisualValidateResult:
    ok: bool
    export_dir: Path
    valid_csv_path: Path
    invalid_csv_path: Path
    xlsx_path: Path | None
    report_path: Path
    report: dict[str, Any]
    records: list[StoryVisualRecord] = field(default_factory=list)
    message: str = ""
    human_dir: Path | None = None
    human_xlsx_path: Path | None = None
    human_sync_error: str = ""


@dataclass
class SiteVisualRetryResult:
    ok: bool
    retried: int
    skipped_max_retries: int
    gemini_message: str
    validate_after: SiteVisualValidateResult | None = None
    message: str = ""
    status: str = "ok"  # ok | dry_run | no_candidates | infrastructure_error | retry_failed
    exit_reason: str = ""
    selected_gemini_profile: int | None = None
    preflight_status: str = ""
    preflight_reason: str = ""
    browser_launch_error: str = ""
    invalid_total_before_retry: int = 0
    retry_candidates: int = 0
    retry_succeeded: int = 0
    retry_failed: int = 0
    retry_skipped: int = 0
    final_valid_count: int = 0
    final_invalid_count: int = 0
    report_path: Path | None = None
    preflight_report: dict[str, Any] = field(default_factory=dict)


def _read_text(path: Path) -> str:
    if not path.is_file():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def _read_json(path: Path) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return payload if isinstance(payload, dict) else None


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_retry_state(export_dir: Path) -> dict[str, int]:
    path = export_dir / RETRY_STATE_FILE
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    out: dict[str, int] = {}
    for k, v in data.items():
        try:
            out[str(k)] = int(v)
        except (TypeError, ValueError):
            continue
    return out


def _save_retry_state(export_dir: Path, state: dict[str, int]) -> None:
    _write_json(export_dir / RETRY_STATE_FILE, state)


def resolve_human_visual_dir(runs_root: Path | None) -> Path | None:
    """
    По runs_root (e.g. <launch>/10_Временные_файлы/legacy/runs/site/<id>-a)
    вернуть <launch>/02_Сайт/03_Визуал_для_сайта.
    Если launch не определяется — None (вызывающий код должен ошибиться явно).
    """
    if runs_root is None:
        return None
    p = Path(runs_root).resolve()
    for ancestor in [p] + list(p.parents):
        if (ancestor / HUMAN_LAUNCH_ROOT_MARKER).is_dir():
            launch = ancestor
            return launch.joinpath(*HUMAN_LAUNCH_VISUAL_REL)
    return None


def sync_visual_artifacts_to_human(
    *,
    export_dir: Path,
    human_dir: Path,
) -> dict[str, str]:
    """Копирует CSV/XLSX/JSON из export_dir в human_dir. Перезаписывает существующие."""
    human_dir.mkdir(parents=True, exist_ok=True)
    copied: dict[str, str] = {}
    pairs = (
        ("visual_prompts.xlsx", HUMAN_VISUAL_XLSX_NAME),
        ("visual_prompts.csv", HUMAN_VISUAL_CSV_NAME),
        ("visual_prompts_invalid.csv", HUMAN_VISUAL_INVALID_CSV_NAME),
        ("visual_prompts_build_report.json", HUMAN_VISUAL_REPORT_NAME),
        ("visual_prompts_retry_report.json", HUMAN_VISUAL_RETRY_REPORT_NAME),
    )
    for src_name, dst_name in pairs:
        src = export_dir / src_name
        if not src.is_file():
            continue
        dst = human_dir / dst_name
        shutil.copy2(src, dst)
        copied[dst_name] = str(dst)
    return copied


def _find_cleaned_text(story_workspace: Path, canonical: str) -> Path | None:
    for name in (f"{canonical}__M.txt", f"{canonical}__F.txt", f"{canonical}__U.txt", "cleaned_story.txt"):
        p = story_workspace / name
        if p.is_file() and p.stat().st_size > 0:
            return p
    voice = sorted(story_workspace.glob("*__[MFU].txt"), key=lambda x: x.name.lower())
    if voice:
        return voice[0]
    txts = [p for p in story_workspace.glob("*.txt") if p.name.lower() != "info.txt"]
    if txts:
        return sorted(txts, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return None


def collect_story_visual_records(
    *,
    runs_stories_dir: Path,
    output_site_dir: Path,
    export_dir: Path | None = None,
) -> list[StoryVisualRecord]:
    retry_state = _load_retry_state(export_dir) if export_dir else {}
    records: list[StoryVisualRecord] = []

    if not runs_stories_dir.is_dir():
        return records

    for story_dir in sorted([p for p in runs_stories_dir.iterdir() if p.is_dir()], key=lambda x: x.name.lower()):
        pipeline = story_dir / "_pipeline"
        mapping = _read_json(pipeline / "mapping.json") if pipeline.is_dir() else None
        canonical = str((mapping or {}).get("canonical_basename") or story_dir.name).strip()
        story_id = str((mapping or {}).get("story_id") or story_dir.name).strip()

        raw_path = pipeline / "site_info_raw.txt" if pipeline.is_dir() else None
        if raw_path is not None and not raw_path.is_file():
            raw_path = None

        raw_text = _read_text(raw_path) if raw_path else ""
        site_info_json = _read_json(pipeline / "site_info.json") if pipeline.is_dir() else None
        output_story = output_site_dir / canonical
        info_text = _read_text(output_story / "info.txt") if output_story.is_dir() else _read_text(story_dir / "info.txt")

        visual_full, extraction_source, status, failure_reason, excerpt = validate_story_visual_from_raw(
            canonical_basename=canonical,
            raw_path=raw_path,
            raw_text=raw_text,
            fallback_json=site_info_json,
            fallback_info_text=info_text,
        )

        preview = make_visual_prompt_preview(visual_full) if visual_full else ""
        final_prompt = f"{TECHNICAL_PROMPT}; {visual_full}" if status == "ok" and visual_full else ""

        records.append(
            StoryVisualRecord(
                canonical_basename=canonical,
                story_id=story_id,
                story_workspace_path=story_dir.resolve(),
                output_story_dir=output_story.resolve() if output_story.is_dir() else output_story,
                raw_path=raw_path.resolve() if raw_path and raw_path.is_file() else None,
                raw_text=raw_text,
                site_info_json=site_info_json,
                info_text=info_text,
                visual_prompt_full=visual_full if status == "ok" else "",
                visual_prompt_preview=preview,
                visual_prompt_status=status,
                failure_reason=failure_reason,
                extraction_source=extraction_source,
                final_prompt=final_prompt,
                retry_count=retry_state.get(story_id, 0),
                raw_excerpt_preview=excerpt,
            )
        )
    return records


def write_visual_prompt_tables(
    records: list[StoryVisualRecord],
    export_dir: Path,
    *,
    summary_extra: dict[str, Any] | None = None,
) -> tuple[Path, Path, Path, Path, dict[str, Any]]:
    export_dir.mkdir(parents=True, exist_ok=True)
    counters = build_validation_report_counters()

    valid_rows = [r for r in records if r.is_valid]
    invalid_rows = [r for r in records if not r.is_valid]

    for row in records:
        bump_report_counter(counters, row.failure_reason, row.visual_prompt_status)

    valid_csv = export_dir / "visual_prompts.csv"
    invalid_csv = export_dir / "visual_prompts_invalid.csv"
    report_path = export_dir / "visual_prompts_build_report.json"
    xlsx_path = export_dir / "visual_prompts.xlsx"

    with valid_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(VALID_CSV_COLUMNS)
        for row in valid_rows:
            writer.writerow(
                [
                    row.canonical_basename,
                    row.story_id,
                    TECHNICAL_PROMPT,
                    row.visual_prompt_full,
                    row.visual_prompt_preview,
                    row.final_prompt,
                    row.visual_prompt_status,
                    row.extraction_source,
                    str(row.story_workspace_path),
                ]
            )

    with invalid_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(INVALID_CSV_COLUMNS)
        for row in invalid_rows:
            writer.writerow(
                [
                    row.canonical_basename,
                    row.story_id,
                    row.visual_prompt_status,
                    row.failure_reason,
                    row.extraction_source,
                    str(row.raw_path) if row.raw_path else "",
                    str(row.story_workspace_path),
                    row.retry_count,
                    row.raw_excerpt_preview,
                ]
            )

    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl required for visual_prompts.xlsx. Install: pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws_valid = wb.active
    ws_valid.title = "valid_prompts"
    ws_valid.append(VALID_CSV_COLUMNS)
    for row in valid_rows:
        ws_valid.append(
            [
                row.canonical_basename,
                row.story_id,
                TECHNICAL_PROMPT,
                row.visual_prompt_full,
                row.visual_prompt_preview,
                row.final_prompt,
                row.visual_prompt_status,
                row.extraction_source,
                str(row.story_workspace_path),
            ]
        )

    ws_invalid = wb.create_sheet("invalid_prompts")
    ws_invalid.append(INVALID_CSV_COLUMNS)
    for row in invalid_rows:
        ws_invalid.append(
            [
                row.canonical_basename,
                row.story_id,
                row.visual_prompt_status,
                row.failure_reason,
                row.extraction_source,
                str(row.raw_path) if row.raw_path else "",
                str(row.story_workspace_path),
                row.retry_count,
                row.raw_excerpt_preview,
            ]
        )

    ws_summary = wb.create_sheet("summary")
    summary_rows: list[tuple[str, Any]] = [
        ("metric", "value"),
        *[(k, counters[k]) for k in counters],
        ("valid_rows", len(valid_rows)),
        ("invalid_rows", len(invalid_rows)),
    ]
    if summary_extra:
        for k, v in summary_extra.items():
            if isinstance(v, (dict, list)):
                summary_rows.append((str(k), json.dumps(v, ensure_ascii=False)))
            else:
                summary_rows.append((str(k), v))
    for pair in summary_rows:
        ws_summary.append(list(pair))

    wb.save(xlsx_path)

    report: dict[str, Any] = {
        **counters,
        "valid_csv_path": str(valid_csv),
        "invalid_csv_path": str(invalid_csv),
        "xlsx_path": str(xlsx_path),
        "valid_rows": len(valid_rows),
        "invalid_rows": len(invalid_rows),
    }
    _write_json(report_path, report)

    return valid_csv, invalid_csv, xlsx_path, report_path, report


def run_validate_site_info_visuals(
    *,
    runs_stories_dir: Path,
    output_site_dir: Path,
    export_dir: Path,
    runs_root: Path | None = None,
    sync_to_human: bool = True,
    require_human_dir: bool = False,
    summary_extra: dict[str, Any] | None = None,
) -> SiteVisualValidateResult:
    records = collect_story_visual_records(
        runs_stories_dir=runs_stories_dir,
        output_site_dir=output_site_dir,
        export_dir=export_dir,
    )
    try:
        valid_csv, invalid_csv, xlsx_path, report_path, report = write_visual_prompt_tables(
            records,
            export_dir,
            summary_extra=summary_extra,
        )
    except RuntimeError as exc:
        return SiteVisualValidateResult(
            ok=False,
            export_dir=export_dir,
            valid_csv_path=export_dir / "visual_prompts.csv",
            invalid_csv_path=export_dir / "visual_prompts_invalid.csv",
            xlsx_path=None,
            report_path=export_dir / "visual_prompts_build_report.json",
            report={},
            records=records,
            message=str(exc),
        )

    invalid_n = int(report.get("invalid_prompts", 0))
    valid_n = int(report.get("valid_prompts", 0))
    msg = f"validate_site_info_visuals: valid={valid_n} invalid={invalid_n}"
    print(f"[visual-validate] {msg}", flush=True)
    if invalid_n > 0:
        print(
            f"[visual-validate] invalid file: {invalid_csv} (retry via site-info-visual retry)",
            flush=True,
        )

    # ---- Human Excel sync ----
    candidate_runs_root = runs_root if runs_root is not None else export_dir.parent
    human_dir = resolve_human_visual_dir(candidate_runs_root)
    human_xlsx_path: Path | None = None
    human_sync_error = ""
    if sync_to_human:
        if human_dir is None:
            human_sync_error = (
                "human_visual_prompts_xlsx_path_not_found: не удалось определить launch "
                f"(ожидался каталог '<launch>/{HUMAN_LAUNCH_ROOT_MARKER}'). runs_root={candidate_runs_root}"
            )
            print(f"[visual-validate] {human_sync_error}", flush=True)
        else:
            try:
                copied = sync_visual_artifacts_to_human(export_dir=export_dir, human_dir=human_dir)
                human_xlsx_path = human_dir / HUMAN_VISUAL_XLSX_NAME if (human_dir / HUMAN_VISUAL_XLSX_NAME).is_file() else None
                print(
                    f"[visual-validate] human-sync OK: {human_dir} ({len(copied)} files)",
                    flush=True,
                )
            except Exception as exc:
                human_sync_error = f"human_sync_failed: {exc}"
                print(f"[visual-validate] {human_sync_error}", flush=True)

    ok_overall = True
    final_msg = msg
    if require_human_dir and (human_dir is None or human_sync_error):
        ok_overall = False
        final_msg = human_sync_error or "human_visual_prompts_xlsx_path_not_found"

    return SiteVisualValidateResult(
        ok=ok_overall,
        export_dir=export_dir,
        valid_csv_path=valid_csv,
        invalid_csv_path=invalid_csv,
        xlsx_path=xlsx_path,
        report_path=report_path,
        report=report,
        records=records,
        message=final_msg,
        human_dir=human_dir,
        human_xlsx_path=human_xlsx_path,
        human_sync_error=human_sync_error,
    )


def _read_invalid_csv_rows(invalid_csv: Path) -> list[dict[str, str]]:
    if not invalid_csv.is_file():
        return []
    with invalid_csv.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f, delimiter=";"))


def _apply_gemini_info_to_workspace(
    *,
    story_workspace: Path,
    output_story_dir: Path,
    raw_text: str,
    parse_site_info_result,
    render_legacy_info,
) -> None:
    pipeline = story_workspace / "_pipeline"
    pipeline.mkdir(parents=True, exist_ok=True)
    mapping = _read_json(pipeline / "mapping.json") or {}
    story_id = str(mapping.get("story_id") or story_workspace.name)
    canonical = str(mapping.get("canonical_basename") or story_workspace.name)

    (pipeline / "site_info_raw.txt").write_text(raw_text, encoding="utf-8")
    site_info = parse_site_info_result(story_id, canonical, raw_text)
    _write_json(pipeline / "site_info.json", site_info)
    legacy_info = render_legacy_info(site_info)
    (story_workspace / "info.txt").write_text(legacy_info, encoding="utf-8")
    if output_story_dir.parent.exists():
        output_story_dir.mkdir(parents=True, exist_ok=True)
        (output_story_dir / "info.txt").write_text(legacy_info, encoding="utf-8")


def _write_retry_report(export_dir: Path, payload: dict[str, Any]) -> Path:
    path = export_dir / "visual_prompts_retry_report.json"
    _write_json(path, payload)
    return path


def _read_gemini_worker_log_tail(logs_dir: Path, stage_key: str, *, lines: int = 120) -> str:
    log_path = logs_dir / f"{stage_key}_worker_1.log"
    if not log_path.is_file():
        return ""
    try:
        return "\n".join(log_path.read_text(encoding="utf-8", errors="ignore").splitlines()[-lines:])
    except OSError:
        return ""


def _detect_browser_failure(msg: str, *, extra_context: str = "") -> str:
    """Если сообщение/лог Gemini gate выглядит как браузерный сбой, вернуть код причины."""
    blob = f"{msg}\n{extra_context}".strip()
    if not blob:
        return ""
    low = blob.lower()
    triggers = [
        ("processsingleton", "browser_profile_locked"),
        ("profile is already in use", "browser_profile_locked"),
        ("singletonlock", "browser_profile_locked"),
        ("target page, context or browser has been closed", "browser_context_closed"),
        ("context or browser has been closed", "browser_context_closed"),
        ("startup_failed_after_retries", "browser_context_closed"),
        ("executable doesn't exist", "playwright_binary_missing"),
        ("browser has disconnected", "browser_disconnected"),
        ("net::err_", "browser_network_error"),
        ("err_name_not_resolved", "browser_network_error"),
        ("ui не готов", "gemini_ui_not_ready"),
        ("не удалось открыть сессию gemini", "gemini_ui_not_ready"),
        ("не удалось открыть url гем", "gemini_ui_not_ready"),
    ]
    for needle, code in triggers:
        if needle in low:
            return code
    if "legacy gemini gate failed" in low and extra_context.strip():
        return "gemini_runner_failed"
    return ""


def run_retry_invalid_site_info_visuals(
    *,
    config: OrchestratorConfig,
    runs_root: Path,
    output_site_dir: Path,
    export_dir: Path,
    gemini_registry_path: Path,
    gemini_info_stage_key: str = "site_info_builder",
    gemini_workers: int = 1,
    max_retry_attempts: int = 2,
    execute: bool = False,
    gemini_target_active_workers: int = 3,
    gemini_profiles_total: int = 5,
    gemini_max_restarts_per_profile: int = 3,
    gemini_profile_cooldown_seconds: float = 900.0,
    gemini_supervised_workers: bool = True,
    profile_index: int | None = None,
    auto_profile: bool = False,
) -> SiteVisualRetryResult:
    from orchestrator.phase_a import (
        _build_gemini_input,
        _parse_site_info_result,
        _render_legacy_info,
        _run_legacy_gemini_gate,
    )
    from orchestrator.site_visual_profile_preflight import (
        preflight_to_dict,
        run_profile_preflight,
    )

    runs_stories_dir = runs_root / "stories"
    invalid_csv = export_dir / "visual_prompts_invalid.csv"
    invalid_rows = _read_invalid_csv_rows(invalid_csv)
    invalid_total = len(invalid_rows)

    def _build_report(result: SiteVisualRetryResult) -> dict[str, Any]:
        return {
            "status": result.status,
            "ok": result.ok,
            "exit_reason": result.exit_reason or result.message,
            "invalid_total_before_retry": result.invalid_total_before_retry,
            "retry_candidates": result.retry_candidates,
            "retried": result.retried,
            "retry_succeeded": result.retry_succeeded,
            "retry_failed": result.retry_failed,
            "retry_skipped": result.retry_skipped,
            "skipped_max_retries": result.skipped_max_retries,
            "selected_gemini_profile": result.selected_gemini_profile,
            "profile_preflight_status": result.preflight_status,
            "profile_preflight_reason": result.preflight_reason,
            "browser_launch_error": result.browser_launch_error,
            "final_valid_count": result.final_valid_count,
            "final_invalid_count": result.final_invalid_count,
            "gemini_message": result.gemini_message,
            "message": result.message,
            "auto_profile": auto_profile,
            "requested_profile_index": profile_index,
            "max_retry_attempts": max_retry_attempts,
            "preflight_report": result.preflight_report,
        }

    def _finalize(result: SiteVisualRetryResult) -> SiteVisualRetryResult:
        # Если retry не выполнил полный цикл (dry-run / no_candidates / infrastructure_error /
        # retry_failed), пересобираем таблицы и human-копию с явным retry-summary, чтобы
        # пользователь видел, почему ничего не изменилось.
        if result.validate_after is None:
            try:
                result.validate_after = run_validate_site_info_visuals(
                    runs_stories_dir=runs_stories_dir,
                    output_site_dir=output_site_dir,
                    export_dir=export_dir,
                    runs_root=runs_root,
                    summary_extra={
                        "retry_status": result.status,
                        "retry_exit_reason": result.exit_reason,
                        "selected_gemini_profile": (
                            f"user_data_{result.selected_gemini_profile}"
                            if result.selected_gemini_profile is not None
                            else "n/a"
                        ),
                        "preflight_status": result.preflight_status,
                        "preflight_reason": result.preflight_reason,
                        "browser_launch_error": result.browser_launch_error[:300] if result.browser_launch_error else "",
                        "invalid_total_before_retry": result.invalid_total_before_retry,
                        "retry_candidates": result.retry_candidates,
                        "retry_skipped_max_retries": result.retry_skipped,
                    },
                )
            except Exception as exc:
                print(f"[visual-retry] post-finalize validate failed: {exc}", flush=True)
        rep: dict[str, Any] = {}
        if result.validate_after is not None:
            rep = result.validate_after.report or {}
        else:
            br = export_dir / "visual_prompts_build_report.json"
            if br.is_file():
                try:
                    loaded = json.loads(br.read_text(encoding="utf-8"))
                    if isinstance(loaded, dict):
                        rep = loaded
                except Exception:
                    rep = {}
        if rep:
            result.final_valid_count = int(rep.get("valid_prompts", rep.get("valid_rows", 0)))
            result.final_invalid_count = int(rep.get("invalid_prompts", rep.get("invalid_rows", 0)))
        report = _build_report(result)
        result.report_path = _write_retry_report(export_dir, report)
        return result

    retry_state = _load_retry_state(export_dir)
    pending: list[dict[str, str]] = []
    skipped_max = 0

    for row in invalid_rows:
        reason = (row.get("failure_reason") or "").strip()
        if reason not in RETRYABLE_FAILURE_REASONS:
            continue
        story_id = (row.get("story_id") or "").strip()
        if not story_id:
            continue
        count = retry_state.get(story_id, 0)
        if count >= max_retry_attempts:
            skipped_max += 1
            continue
        pending.append(row)

    if not pending:
        msg = "retry: нет историй для повторного site_info_builder (или исчерпан max_retry_attempts)"
        print(f"[visual-retry] {msg}", flush=True)
        validate_after = run_validate_site_info_visuals(
            runs_stories_dir=runs_stories_dir,
            output_site_dir=output_site_dir,
            export_dir=export_dir,
            runs_root=runs_root,
            summary_extra={
                "retry_status": "no_candidates",
                "retry_candidates": 0,
                "retry_skipped_max_retries": skipped_max,
                "invalid_total_before_retry": invalid_total,
            },
        )
        return _finalize(SiteVisualRetryResult(
            ok=True,
            retried=0,
            skipped_max_retries=skipped_max,
            gemini_message=msg,
            validate_after=validate_after,
            message=msg,
            status="no_candidates",
            exit_reason="no_retryable_candidates",
            invalid_total_before_retry=invalid_total,
            retry_candidates=0,
            retry_skipped=skipped_max,
        ))

    if not execute:
        msg = (
            f"retry dry-run: {len(pending)} stories queued "
            f"(add --execute to run Gemini {gemini_info_stage_key})"
        )
        print(f"[visual-retry] {msg}", flush=True)
        return _finalize(SiteVisualRetryResult(
            ok=True,
            retried=0,
            skipped_max_retries=skipped_max,
            gemini_message=msg,
            message=msg,
            status="dry_run",
            exit_reason="dry_run",
            invalid_total_before_retry=invalid_total,
            retry_candidates=len(pending),
            retry_skipped=skipped_max,
        ))

    # ---- Browser/profile preflight ----
    preflight = run_profile_preflight(
        config=config,
        registry_path=gemini_registry_path,
        stage_key=gemini_info_stage_key,
        profiles_total=max(1, int(gemini_profiles_total)),
        requested_profile_index=profile_index,
        auto_profile=bool(auto_profile),
    )
    preflight_dict = preflight_to_dict(preflight)
    print(
        f"[visual-retry][preflight] status={preflight.preflight_status} "
        f"selected={preflight.selected_profile_index} reason={preflight.reason}",
        flush=True,
    )
    for p in preflight.profiles:
        marker = "READY" if p.is_ready else ("LOCKED" if p.is_locked else "NOT_READY")
        print(
            f"[visual-retry][preflight] profile {p.profile_index} {marker} "
            f"email={p.email or 'n/a'} url_for_stage={p.registry_url_for_stage} "
            f"locks={p.lock_files_present or '-'} chrome_pids={p.chrome_pids or '-'} "
            f"reasons={p.reasons or '-'}",
            flush=True,
        )

    if not preflight.ok:
        reason_code = (
            "browser_profile_locked"
            if any(p.is_locked for p in preflight.profiles)
            else (
                "registry_missing"
                if preflight.preflight_status == "registry_missing"
                else "no_ready_profile"
            )
        )
        msg = (
            f"infrastructure_error: {reason_code}; preflight={preflight.preflight_status}; "
            f"reason={preflight.reason}"
        )
        print(f"[visual-retry][infrastructure_error] {msg}", flush=True)
        return _finalize(SiteVisualRetryResult(
            ok=False,
            retried=0,
            skipped_max_retries=skipped_max,
            gemini_message=msg,
            message=msg,
            status="infrastructure_error",
            exit_reason=reason_code,
            selected_gemini_profile=preflight.selected_profile_index,
            preflight_status=preflight.preflight_status,
            preflight_reason=preflight.reason,
            browser_launch_error=msg,
            invalid_total_before_retry=invalid_total,
            retry_candidates=len(pending),
            retry_skipped=skipped_max,
            preflight_report=preflight_dict,
        ))

    profile_try_order: list[int] = []
    if auto_profile:
        profile_try_order = [p.profile_index for p in preflight.profiles if p.is_ready]
    elif profile_index is not None:
        profile_try_order = [int(profile_index)]
    else:
        profile_try_order = [int(preflight.selected_profile_index or 0)]
    if not profile_try_order:
        profile_try_order = [0]

    # ---- Подготовка ввода Gemini ----
    cleaned_paths: list[Path] = []
    story_meta: list[tuple[Path, Path, Path]] = []

    for row in pending:
        ws_path = Path((row.get("story_workspace_path") or "").strip())
        if not ws_path.is_dir():
            story_id = row.get("story_id", "")
            alt = runs_stories_dir / story_id
            if alt.is_dir():
                ws_path = alt
        if not ws_path.is_dir():
            continue
        canonical = (row.get("canonical_basename") or ws_path.name).strip()
        cleaned = _find_cleaned_text(ws_path, canonical)
        if cleaned is None:
            print(f"[visual-retry] skip {canonical}: no cleaned text", flush=True)
            continue
        cleaned_paths.append(cleaned)
        out_story = output_site_dir / canonical
        story_meta.append((ws_path, out_story, cleaned))

    if not cleaned_paths:
        return _finalize(SiteVisualRetryResult(
            ok=False,
            retried=0,
            skipped_max_retries=skipped_max,
            gemini_message="no cleaned files for retry queue",
            message="no cleaned files for retry queue",
            status="retry_failed",
            exit_reason="no_cleaned_text",
            selected_gemini_profile=selected_profile,
            preflight_status=preflight.preflight_status,
            preflight_reason=preflight.reason,
            invalid_total_before_retry=invalid_total,
            retry_candidates=len(pending),
            retry_skipped=skipped_max,
            preflight_report=preflight_dict,
        ))

    info_root = runs_root / "gemini_info_retry"
    if info_root.exists():
        shutil.rmtree(info_root, ignore_errors=True)
    info_input_root, info_mapping = _build_gemini_input(cleaned_paths, runs_stories_dir, info_root)

    for item in info_mapping:
        gemini_dir = Path(item["gemini_story_dir"])
        info_txt = gemini_dir / "info.txt"
        if info_txt.is_file():
            info_txt.unlink()

    logs_dir = runs_root / "logs" / "visual_retry"
    ok_info = False
    info_msg = ""
    selected_profile = profile_try_order[0]
    last_log_tail = ""
    for attempt_idx, prof_idx in enumerate(profile_try_order):
        selected_profile = int(prof_idx)
        if attempt_idx > 0:
            print(
                f"[visual-retry] fallback profile user_data_{selected_profile} "
                f"(attempt {attempt_idx + 1}/{len(profile_try_order)})",
                flush=True,
            )
        print(
            f"[visual-retry] launching Gemini {gemini_info_stage_key} on profile user_data_{selected_profile} "
            f"(workers={gemini_workers}, queue={len(info_mapping)})",
            flush=True,
        )
        ok_info, info_msg = _run_legacy_gemini_gate(
            config,
            info_input_root,
            gemini_registry_path,
            gemini_info_stage_key,
            gemini_workers,
            logs_dir,
            events_file=config.events_file,
            run_id=runs_root.name,
            target_active_workers=max(1, min(5, int(gemini_target_active_workers))),
            pending_count=len(info_mapping),
            profiles_total=max(1, min(5, int(gemini_profiles_total))),
            max_restarts_per_profile=max(1, int(gemini_max_restarts_per_profile)),
            profile_cooldown_seconds=float(gemini_profile_cooldown_seconds),
            supervised_workers=False,
            profile_index_override=selected_profile,
        )
        last_log_tail = _read_gemini_worker_log_tail(logs_dir, gemini_info_stage_key)
        if ok_info:
            if attempt_idx > 0:
                preflight_dict["launch_fallback_profile"] = selected_profile
            break
        if not auto_profile or attempt_idx >= len(profile_try_order) - 1:
            break

    if not ok_info:
        infra_code = _detect_browser_failure(info_msg, extra_context=last_log_tail)
        is_infra = bool(infra_code)
        browser_err = (last_log_tail or info_msg)[:4000] if is_infra else ""
        fail_msg = info_msg
        if is_infra and last_log_tail:
            fail_msg = f"{info_msg}\n--- worker log tail ---\n{last_log_tail[-1500:]}"
        return _finalize(SiteVisualRetryResult(
            ok=False,
            retried=0,
            skipped_max_retries=skipped_max,
            gemini_message=fail_msg,
            message=fail_msg,
            status="infrastructure_error" if is_infra else "retry_failed",
            exit_reason=infra_code or "gemini_gate_failed",
            selected_gemini_profile=selected_profile,
            preflight_status=preflight.preflight_status,
            preflight_reason=preflight.reason,
            browser_launch_error=browser_err,
            invalid_total_before_retry=invalid_total,
            retry_candidates=len(pending),
            retry_failed=len(pending),
            retry_skipped=skipped_max,
            preflight_report=preflight_dict,
        ))

    applied = 0
    mapping_by_cleaned: dict[str, dict[str, str]] = {}
    for item in info_mapping:
        mapping_by_cleaned[str(Path(item["source_path"]).resolve())] = item

    for ws_path, out_story, cleaned in story_meta:
        item = mapping_by_cleaned.get(str(cleaned.resolve()))
        if not item:
            continue
        src_info = Path(item["gemini_story_dir"]) / "info.txt"
        if not src_info.is_file():
            continue
        raw_text = src_info.read_text(encoding="utf-8", errors="ignore")
        _apply_gemini_info_to_workspace(
            story_workspace=ws_path,
            output_story_dir=out_story,
            raw_text=raw_text,
            parse_site_info_result=_parse_site_info_result,
            render_legacy_info=_render_legacy_info,
        )
        pipeline = ws_path / "_pipeline"
        mapping = _read_json(pipeline / "mapping.json") or {}
        sid = str(mapping.get("story_id") or ws_path.name)
        retry_state[sid] = retry_state.get(sid, 0) + 1
        applied += 1

    _save_retry_state(export_dir, retry_state)

    # Запомним множество story_id, которые именно сейчас пытались переретраить.
    retried_story_ids: set[str] = set()
    for row in pending:
        sid = (row.get("story_id") or "").strip()
        if sid:
            retried_story_ids.add(sid)

    validate_after = run_validate_site_info_visuals(
        runs_stories_dir=runs_stories_dir,
        output_site_dir=output_site_dir,
        export_dir=export_dir,
        runs_root=runs_root,
        summary_extra={
            "retry_status": "ok",
            "retry_candidates": len(pending),
            "retry_applied": applied,
            "retry_skipped_max_retries": skipped_max,
            "selected_gemini_profile": f"user_data_{selected_profile}",
            "preflight_status": preflight.preflight_status,
            "invalid_total_before_retry": invalid_total,
        },
    )

    # Точный подсчёт succeeded/failed по story_id из retried set.
    succeeded = 0
    failed_after_retry = 0
    for rec in validate_after.records:
        if rec.story_id in retried_story_ids:
            if rec.is_valid:
                succeeded += 1
            else:
                failed_after_retry += 1
    # Если истории не нашлись в результатах (например, удалили) — считаем как failed.
    missing = len(retried_story_ids) - succeeded - failed_after_retry
    failed_in_gemini = max(0, failed_after_retry + missing)

    valid_after = int(validate_after.report.get("valid_prompts", validate_after.report.get("valid_rows", 0)))
    invalid_after = int(validate_after.report.get("invalid_prompts", validate_after.report.get("invalid_rows", 0)))
    msg = (
        f"retry done: applied={applied} succeeded={succeeded} failed={failed_in_gemini} "
        f"skipped_max_retries={skipped_max} "
        f"valid_after={valid_after} invalid_after={invalid_after} profile=user_data_{selected_profile}"
    )
    print(f"[visual-retry] {msg}", flush=True)
    return _finalize(SiteVisualRetryResult(
        ok=True,
        retried=applied,
        skipped_max_retries=skipped_max,
        gemini_message=info_msg,
        validate_after=validate_after,
        message=msg,
        status="ok",
        exit_reason="ok",
        selected_gemini_profile=selected_profile,
        preflight_status=preflight.preflight_status,
        preflight_reason=preflight.reason,
        invalid_total_before_retry=invalid_total,
        retry_candidates=len(pending),
        retry_succeeded=succeeded,
        retry_failed=failed_in_gemini,
        retry_skipped=skipped_max,
        preflight_report=preflight_dict,
    ))


def run_site_info_visual_full_cycle(
    *,
    config: OrchestratorConfig,
    runs_root: Path,
    output_site_dir: Path,
    export_dir: Path,
    gemini_registry_path: Path,
    execute: bool = False,
    max_retry_attempts: int = 2,
    profile_index: int | None = None,
    auto_profile: bool = False,
    **gemini_kwargs: Any,
) -> dict[str, Any]:
    validate_first = run_validate_site_info_visuals(
        runs_stories_dir=runs_root / "stories",
        output_site_dir=output_site_dir,
        export_dir=export_dir,
        runs_root=runs_root,
        summary_extra={"stage": "validate_before_retry"},
    )
    retry = run_retry_invalid_site_info_visuals(
        config=config,
        runs_root=runs_root,
        output_site_dir=output_site_dir,
        export_dir=export_dir,
        gemini_registry_path=gemini_registry_path,
        max_retry_attempts=max_retry_attempts,
        execute=execute,
        profile_index=profile_index,
        auto_profile=auto_profile,
        **gemini_kwargs,
    )
    validate_final = retry.validate_after
    return {
        "ok": validate_first.ok and retry.ok,
        "validate_before": validate_first.report,
        "retry": {
            "status": retry.status,
            "exit_reason": retry.exit_reason,
            "retried": retry.retried,
            "retry_candidates": retry.retry_candidates,
            "retry_succeeded": retry.retry_succeeded,
            "retry_failed": retry.retry_failed,
            "retry_skipped": retry.retry_skipped,
            "skipped_max_retries": retry.skipped_max_retries,
            "selected_gemini_profile": retry.selected_gemini_profile,
            "preflight_status": retry.preflight_status,
            "preflight_reason": retry.preflight_reason,
            "browser_launch_error": retry.browser_launch_error,
            "final_valid_count": retry.final_valid_count,
            "final_invalid_count": retry.final_invalid_count,
            "report_path": str(retry.report_path) if retry.report_path else "",
            "message": retry.message,
        },
        "validate_after": validate_final.report if validate_final is not None else None,
    }


def print_visual_gate_summary(export_dir: Path) -> None:
    report_path = export_dir / "visual_prompts_build_report.json"
    if not report_path.is_file():
        print("[visual-gate] no build report; run validate_site_info_visuals first", flush=True)
        return
    report = json.loads(report_path.read_text(encoding="utf-8"))
    valid_n = int(report.get("valid_prompts", report.get("valid_rows", 0)))
    invalid_n = int(report.get("invalid_prompts", report.get("invalid_rows", 0)))
    print(
        f"[visual-gate] valid_count={valid_n} invalid_count={invalid_n} "
        f"valid_csv={report.get('valid_csv_path')} invalid_csv={report.get('invalid_csv_path')}",
        flush=True,
    )
    if invalid_n > 0:
        print("[visual-gate] ComfyUI auto generation uses VALID rows only; invalid stories need retry.", flush=True)
